from admin import admin_only
from models import *
from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, FSInputFile
from aiogram.filters import Command
from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup
from sqlalchemy import func
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
import os

# Import your models and Session
from models import Session, User, Admin, Category, ActiveSubscription, SuspendedSubscription, ReferralData

router_stats = Router()

def get_admin_keyboard():
    return InlineKeyboardMarkup(inline_keyboard=[
    [
        InlineKeyboardButton(text="Отчет по пользователям 👥", callback_data="report_users"),
        InlineKeyboardButton(text="Отчет по категориям 📑", callback_data="report_categories")
    ],
    [
        InlineKeyboardButton(text="Финансовый отчет 💰", callback_data="report_financial"),
        InlineKeyboardButton(text="Отчет по подпискам 📊", callback_data="report_subs")
    ],
    [
        InlineKeyboardButton(text="Полный отчет 📈", callback_data="report_full"),
        InlineKeyboardButton(text="Отчет по рефералам 👥", callback_data="report_referrals")
    ]
])

@router_stats.message(Command("admin"))
@admin_only
async def admin_command(message: Message):       
    await message.answer("📊 Панель администратора - Выберите тип отчета:", reply_markup=get_admin_keyboard())


@router_stats.callback_query(F.data.startswith("report_"))
async def handle_report_callbacks(callback: CallbackQuery):
    await callback.answer()
    report_type = callback.data.split("_")[1]
    
    try:
        # Create temporary file path
        temp_file_path = f"temp_{report_type}_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        await callback.message.answer("Генерация отчета, пожалуйста подождите...")

        if report_type == "users":
            await generate_users_report(temp_file_path)
            caption = "Отчет по пользователям"
        elif report_type == "categories":
            await generate_categories_report(temp_file_path)
            caption = "Отчет по категориям"
        elif report_type == "financial":
            await generate_financial_report(temp_file_path)
            caption = "Финансовый отчет"
        elif report_type == "subs":
            await generate_subscriptions_report(temp_file_path)
            caption = "Отчет по подпискам"
        elif report_type == "referrals":
            await generate_referrals_report(temp_file_path)
            caption = "Отчет по рефералам"
        else:  # full report
            await generate_full_report(temp_file_path)
            caption = "Полный отчет"

        # Send document using FSInputFile
        await callback.message.answer_document(
            document=FSInputFile(temp_file_path),
            caption=f"📊 {caption} - {datetime.now().strftime('%d.%m.%Y')}"  # Changed date format to Russian style
        )
        
        # Clean up temporary file
        if os.path.exists(temp_file_path):
            os.remove(temp_file_path)
        
    except Exception as e:
        await callback.message.answer(f"Error generating report: {str(e)}")
        if os.path.exists(temp_file_path):
            os.remove(temp_file_path)

async def generate_users_report(file_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Анализ пользователей"
   
    with Session() as session:
        users = session.query(User).all()
       
        users_data = []
        for user in users:
            active_subs = session.query(func.count(ActiveSubscription.id))\
                .filter(ActiveSubscription.user_id == user.id).scalar()
            referred_users = session.query(func.count(User.id))\
                .filter(User.referred_by_id == user.id).scalar()
               
            users_data.append({
                'ID пользователя': user.id,
                'Имя пользователя': user.username,
                'Полное имя': f"{user.first_name or ''} {user.last_name or ''}".strip(),
                'Активные подписки': active_subs,
                'Приглашенные пользователи': referred_users,
                'Всего категорий': len(user.categories)
            })
   
    df = pd.DataFrame(users_data)
   
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)
        if r_idx == 1:
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
   
    wb.save(file_path)

async def generate_categories_report(file_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Анализ категорий"
   
    with Session() as session:
        categories = session.query(Category).all()
       
        categories_data = []
        for category in categories:
            active_subs = session.query(func.count(ActiveSubscription.id))\
                .filter(ActiveSubscription.category_id == category.id).scalar()
            suspended_subs = session.query(func.count(SuspendedSubscription.id))\
                .filter(SuspendedSubscription.category_id == category.id).scalar()
            
            # Get count of trial users
            trial_users = session.query(func.count(UsedTrial.id))\
                .filter(UsedTrial.category_id == category.id).scalar()
            
            # Get count of active trial subscriptions
            active_trials = session.query(func.count(ActiveSubscription.id))\
                .filter(
                    ActiveSubscription.category_id == category.id,
                    ActiveSubscription.subscription_type == 'trial'
                ).scalar()
               
            categories_data.append({
                'Категория': category.name,
                'Пробный период': 'Да' if category.has_3_days_free else 'Нет',
                'Использовали пробный период': trial_users,
                'Активные пробные': active_trials,
                'Месячная цена': category.price_monthly,
                'Квартальная цена': category.price_quarterly,
                'Полугодовая цена': category.price_half_yearly,
                'Годовая цена': category.price_yearly,
                'Активные пользователи': len(category.users),
                'Активные подписки': active_subs - active_trials,  # Исключаем пробные
                'Приостановленные подписки': suspended_subs,
                'Ключевые слова': category.keywords
            })
   
    df = pd.DataFrame(categories_data)
    
    # Adjust column widths based on content
    max_lengths = {column: max(df[column].astype(str).apply(len).max(), len(column)) + 2 
                  for column in df.columns}
   
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)
        if r_idx == 1:  # Header row
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            
            # Set column widths
            for idx, column in enumerate(df.columns, 1):
                ws.column_dimensions[get_column_letter(idx)].width = max_lengths[column]
    
    # Add some basic formatting for better readability
    for row in ws.iter_rows(min_row=2):  # Starting from second row
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'  # Format numbers with thousand separator
            cell.alignment = Alignment(horizontal='left')  # Left align all cells
   
    wb.save(file_path)

async def generate_financial_report(file_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Финансовый анализ"
   
    with Session() as session:
        categories = session.query(Category).all()
       
        financial_data = []
        total_revenue = 0
       
        for category in categories:
            monthly_subs = session.query(func.count(ActiveSubscription.id))\
                .filter(ActiveSubscription.category_id == category.id)\
                .filter(ActiveSubscription.subscription_type == 'monthly').scalar()
            quarterly_subs = session.query(func.count(ActiveSubscription.id))\
                .filter(ActiveSubscription.category_id == category.id)\
                .filter(ActiveSubscription.subscription_type == 'quarterly').scalar()
            yearly_subs = session.query(func.count(ActiveSubscription.id))\
                .filter(ActiveSubscription.category_id == category.id)\
                .filter(ActiveSubscription.subscription_type == 'yearly').scalar()
               
            monthly_revenue = monthly_subs * category.price_monthly
            quarterly_revenue = quarterly_subs * category.price_quarterly
            yearly_revenue = yearly_subs * category.price_yearly
            category_total = monthly_revenue + quarterly_revenue + yearly_revenue
            total_revenue += category_total
           
            financial_data.append({
                'Категория': category.name,
                'Месячные подписчики': monthly_subs,
                'Месячный доход': monthly_revenue,
                'Квартальные подписчики': quarterly_subs,
                'Квартальный доход': quarterly_revenue,
                'Годовые подписчики': yearly_subs,
                'Годовой доход': yearly_revenue,
                'Общий доход': category_total
            })
   
    df = pd.DataFrame(financial_data)
    df.loc[len(df)] = ['ИТОГО',
                       df['Месячные подписчики'].sum(),
                       df['Месячный доход'].sum(),
                       df['Квартальные подписчики'].sum(),
                       df['Квартальный доход'].sum(),
                       df['Годовые подписчики'].sum(),
                       df['Годовой доход'].sum(),
                       total_revenue]
   
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)
        if r_idx == 1:
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
   
    wb.save(file_path)

async def generate_subscriptions_report(file_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Анализ подписок"
   
    with Session() as session:
        active_subs = session.query(ActiveSubscription).all()
       
        subs_data = []
        for sub in active_subs:
            subs_data.append({
                'Пользователь': f"{sub.user.username or ''} ({sub.user.id})",
                'Категория': sub.category.name,
                'Тип': 'Месячная' if sub.subscription_type == 'monthly' 
                      else 'Квартальная' if sub.subscription_type == 'quarterly'
                      else 'Годовая',
                'Дата начала': sub.start_date.strftime('%d.%m.%Y'),
                'Дата окончания': sub.end_date.strftime('%d.%m.%Y'),
                'Осталось дней': (sub.end_date - datetime.now()).days,
                'Цена': getattr(sub.category, f'price_{sub.subscription_type}')
            })
   
    df = pd.DataFrame(subs_data)
   
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)
        if r_idx == 1:
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
   
    wb.save(file_path)

async def generate_referrals_report(file_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Анализ рефералов"
   
    with Session() as session:
        referral_data = session.query(ReferralData).all()
       
        ref_data = []
        for ref in referral_data:
            if ref.referrals_paid_count > 0:  # Только пользователи с рефералами
                ref_data.append({
                    'Пользователь': f"{ref.user.username or ''} ({ref.user.id})",
                    'Баланс': ref.referral_balance,
                    'Оплаченные рефералы': ref.referrals_paid_count,
                    'Доход': ref.cash_income,
                    'Активации': ref.activations_count,
                    'Сумма выплат': ref.payments_sum,
                    'Средняя выплата': ref.payments_sum / ref.payments_count if ref.payments_count else 0
                })
   
    df = pd.DataFrame(ref_data)
   
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)
        if r_idx == 1:
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
   
    wb.save(file_path)

async def generate_full_report(file_path: str):
    wb = Workbook()
    
    # Users Analysis (unchanged)
    ws = wb.active
    ws.title = "Пользователи"
    with Session() as session:
        users = session.query(User).all()
        users_data = []
        for user in users:
            active_subs = session.query(func.count(ActiveSubscription.id))\
                .filter(ActiveSubscription.user_id == user.id).scalar()
            referred_users = session.query(func.count(User.id))\
                .filter(User.referred_by_id == user.id).scalar()
            trial_count = session.query(func.count(UsedTrial.id))\
                .filter(UsedTrial.user_id == user.id).scalar()
            users_data.append({
                'ID пользователя': user.id,
                'Имя пользователя': user.username,
                'Полное имя': f"{user.first_name or ''} {user.last_name or ''}".strip(),
                'Активные подписки': active_subs,
                'Использовано пробных периодов': trial_count,
                'Приглашенные пользователи': referred_users,
                'Всего категорий': len(user.categories)
            })
    
    df_users = pd.DataFrame(users_data)
    for r_idx, row in enumerate(dataframe_to_rows(df_users, index=False, header=True), 1):
        ws.append(row)
        if r_idx == 1:
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    # Categories Analysis
    ws = wb.create_sheet("Категории")
    with Session() as session:
        categories = session.query(Category).all()
        categories_data = []
        for category in categories:
            active_subs = session.query(func.count(ActiveSubscription.id))\
                .filter(ActiveSubscription.category_id == category.id)\
                .filter(ActiveSubscription.subscription_type != 'trial').scalar()
            trial_users = session.query(func.count(UsedTrial.id))\
                .filter(UsedTrial.category_id == category.id).scalar()
            active_trials = session.query(func.count(ActiveSubscription.id))\
                .filter(
                    ActiveSubscription.category_id == category.id,
                    ActiveSubscription.subscription_type == 'trial'
                ).scalar()
            categories_data.append({
                'Категория': category.name,
                'Пробный период': 'Да' if category.has_3_days_free else 'Нет',
                'Использовали пробный период': trial_users,
                'Активные пробные': active_trials,
                'Месячная цена': category.price_monthly,
                'Квартальная цена': category.price_quarterly,
                'Полугодовая цена': category.price_half_yearly,
                'Годовая цена': category.price_yearly,
                'Активные пользователи': len(category.users),
                'Активные подписки': active_subs,
            })
    
    df_categories = pd.DataFrame(categories_data)
    for r_idx, row in enumerate(dataframe_to_rows(df_categories, index=False, header=True), 1):
        ws.append(row)
        if r_idx == 1:
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    # Financial Analysis
    ws = wb.create_sheet("Финансы")
    with Session() as session:
        financial_data = []
        for category in categories:
            monthly_subs = session.query(func.count(ActiveSubscription.id))\
                .filter(ActiveSubscription.category_id == category.id)\
                .filter(ActiveSubscription.subscription_type == 'monthly').scalar()
            quarterly_subs = session.query(func.count(ActiveSubscription.id))\
                .filter(ActiveSubscription.category_id == category.id)\
                .filter(ActiveSubscription.subscription_type == 'quarterly').scalar()
            half_yearly_subs = session.query(func.count(ActiveSubscription.id))\
                .filter(ActiveSubscription.category_id == category.id)\
                .filter(ActiveSubscription.subscription_type == 'half_yearly').scalar()
            yearly_subs = session.query(func.count(ActiveSubscription.id))\
                .filter(ActiveSubscription.category_id == category.id)\
                .filter(ActiveSubscription.subscription_type == 'yearly').scalar()
            
            monthly_revenue = monthly_subs * category.price_monthly
            quarterly_revenue = quarterly_subs * category.price_quarterly
            half_yearly_revenue = half_yearly_subs * category.price_half_yearly
            yearly_revenue = yearly_subs * category.price_yearly
            
            financial_data.append({
                'Категория': category.name,
                'Месячный доход': monthly_revenue,
                'Квартальный доход': quarterly_revenue,
                'Полугодовой доход': half_yearly_revenue,
                'Годовой доход': yearly_revenue,
                'Общий доход': monthly_revenue + quarterly_revenue + half_yearly_revenue + yearly_revenue
            })
    
    df_financial = pd.DataFrame(financial_data)
    df_financial.loc[len(df_financial)] = ['ИТОГО', 
                                         df_financial['Месячный доход'].sum(),
                                         df_financial['Квартальный доход'].sum(),
                                         df_financial['Полугодовой доход'].sum(),
                                         df_financial['Годовой доход'].sum(),
                                         df_financial['Общий доход'].sum()]
    
    for r_idx, row in enumerate(dataframe_to_rows(df_financial, index=False, header=True), 1):
        ws.append(row)
        if r_idx == 1:
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    # Subscriptions Analysis
    ws = wb.create_sheet("Подписки")
    with Session() as session:
        active_subs = session.query(ActiveSubscription).all()
        subs_data = []
        for sub in active_subs:
            # Определяем тип подписки для отображения
            sub_type_display = {
                'trial': 'Пробный период',
                'monthly': 'Месячная',
                'quarterly': 'Квартальная',
                'half_yearly': 'Полугодовая',
                'yearly': 'Годовая'
            }.get(sub.subscription_type, sub.subscription_type)
            
            # Определяем цену (для пробного периода - 0)
            price = 0 if sub.subscription_type == 'trial' else getattr(
                sub.category, f'price_{sub.subscription_type}', 0
            )
            
            subs_data.append({
                'Пользователь': f"{sub.user.username or ''} ({sub.user.id})",
                'Категория': sub.category.name,
                'Тип': sub_type_display,
                'Дата начала': sub.start_date.strftime('%d.%m.%Y'),
                'Дата окончания': sub.end_date.strftime('%d.%m.%Y'),
                'Осталось дней': (sub.end_date - datetime.now()).days,
                'Цена': price
            })
    
    df_subs = pd.DataFrame(subs_data)
    for r_idx, row in enumerate(dataframe_to_rows(df_subs, index=False, header=True), 1):
        ws.append(row)
        if r_idx == 1:
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    # Referrals Analysis (unchanged)
    ws = wb.create_sheet("Рефералы")
    with Session() as session:
        referral_data = session.query(ReferralData).all()
        ref_data = []
        for ref in referral_data:
            if ref.referrals_paid_count > 0:
                ref_data.append({
                    'Пользователь': f"{ref.user.username or ''} ({ref.user.id})",
                    'Баланс': ref.referral_balance,
                    'Оплаченные рефералы': ref.referrals_paid_count,
                    'Доход': ref.cash_income,
                    'Активации': ref.activations_count,
                    'Сумма выплат': ref.payments_sum,
                    'Средняя выплата': ref.payments_sum / ref.payments_count if ref.payments_count else 0
                })
    
    df_refs = pd.DataFrame(ref_data)
    for r_idx, row in enumerate(dataframe_to_rows(df_refs, index=False, header=True), 1):
        ws.append(row)
        if r_idx == 1:
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    # Summary Sheet
    ws = wb.create_sheet("Сводка", 0)
    summary_data = {
        'Показатель': [
            'Всего пользователей',
            'Активные пользователи',
            'Пользователи с пробным периодом',
            'Всего категорий',
            'Всего активных подписок',
            'Активных пробных периодов',
            'Общий месячный доход',
            'Общий квартальный доход',
            'Общий полугодовой доход',
            'Общий годовой доход',
            'Общий доход',
            'Активные рефералы',
            'Общая сумма реферальных выплат'
        ],
        'Значение': [
            len(df_users),
            df_users['Активные подписки'].astype(bool).sum(),
            df_users['Использовано пробных периодов'].astype(bool).sum(),
            len(df_categories),
            len(df_subs),
            len(df_subs[df_subs['Тип'] == 'Пробный период']),
            df_financial['Месячный доход'].sum(),
            df_financial['Квартальный доход'].sum(),
            df_financial['Полугодовой доход'].sum(),
            df_financial['Годовой доход'].sum(),
            df_financial['Общий доход'].sum(),
            len(df_refs),
            df_refs['Сумма выплат'].sum() if len(df_refs) > 0 else 0
        ]
    }
    
    df_summary = pd.DataFrame(summary_data)
    for r_idx, row in enumerate(dataframe_to_rows(df_summary, index=False, header=True), 1):
        ws.append(row)
        if r_idx == 1:
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    # Adjust column widths for all sheets
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    wb.save(file_path)

